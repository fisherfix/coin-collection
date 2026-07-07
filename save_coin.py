#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
save_coin.py — 钱币数据三重存档工具

每次钱币确认后，必须运行本工具保存到：
  1. coin-archive/{ID}.json   — 机器可读，结构化（重建用）
  2. coin-archive/{ID}.md     — 人可读，参考资料
  3. coin-archive/coins.xlsx  — 整库一表（多重备份）

【用法】
  # 保存单枚
  python save_coin.py {ID}
  
  # 列出所有已存档
  python save_coin.py --list
  
  # 从 index.html 重建存档（如果存档丢失但 index.html 完好）
  python save_coin.py --rebuild
"""

import os
import sys
import io
import json
import re
import subprocess
from datetime import datetime

# Windows console 编码
if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

BASE_DIR = r"C:\Users\zhaojingyun\.qclaw\workspace-agent-d8b9b18a\coin-collection"
ARCHIVE_DIR = os.path.join(BASE_DIR, "coin-archive")
INDEX_PATH = os.path.join(BASE_DIR, "index.html")
EXCEL_PATH = os.path.join(ARCHIVE_DIR, "coins.xlsx")

# 汇率配置（与 index.html 中 recalcStats 保持一致）
EXCHANGE_RATES = {
    "CNY": 1.0,
    "EUR": 7.8,
    "GBP": 9.2,
    "USD": 7.2,
}

# 字段定义（顺序与列名）
FIELDS = [
    "id", "category", "title", "subtitle", "era", "denomination",
    "material", "grade", "reference", "description", "obverse", "reverse",
    "price", "price_currency", "price_cny", "seller", "auction", "lot",
    "year", "tags", "image", "image_size", "created_at", "confirmed_at",
    "git_commit",
]


def ensure_archive_dir():
    if not os.path.exists(ARCHIVE_DIR):
        os.makedirs(ARCHIVE_DIR)
        print(f"✓ 已创建存档目录: {ARCHIVE_DIR}")


def parse_price(price_str):
    """解析价格字符串，提取币种和数值"""
    if not price_str or price_str in ("—", "-", ""):
        return None, None, None
    m = re.match(r'([€£$¥]|US\$|EUR|GBP|USD|CNY)\s*([\d,]+(?:\.\d+)?)', price_str.strip())
    if m:
        symbol = m.group(1)
        amount = float(m.group(2).replace(",", ""))
        # 标准化币种码
        symbol_to_code = {
            "€": "EUR", "£": "GBP", "¥": "CNY", "$": "USD",
            "US$": "USD", "EUR": "EUR", "GBP": "GBP", "USD": "USD", "CNY": "CNY",
        }
        code = symbol_to_code.get(symbol, "CNY")
        cny = round(amount * EXCHANGE_RATES.get(code, 1.0))
        return code, amount, cny
    # 默认 CNY
    m = re.match(r'([\d,]+(?:\.\d+)?)', price_str.strip())
    if m:
        amount = float(m.group(1).replace(",", ""))
        return "CNY", amount, int(amount)
    return None, None, None


def parse_provenance(prov_html):
    """解析来源字段，提取卖家、拍卖、Lot、年份"""
    if not prov_html:
        return "", "", "", ""
    # 提取 <strong>...</strong> 作为卖家（取最后一次 = 实际购买方）
    strong_matches = re.findall(r'<strong>([^<]+)</strong>', prov_html)
    seller = strong_matches[-1] if strong_matches else ""

    # 提取年份（取最后一次 = 实际购买年份）
    year_matches = re.findall(r'\b(1[89]\d{2}|20\d{2})\b', prov_html)
    year = year_matches[-1] if year_matches else ""

    # 提取 Lot（取最后一次 = 实际购买 Lot）
    lot_matches = re.findall(r'Lot\s*(\d+)', prov_html, re.IGNORECASE)
    lot = lot_matches[-1] if lot_matches else ""

    # 提取拍卖：优先从最后一次 "<strong>卖家</strong> · 拍卖名" 抓
    auction = ""
    strong_auction_matches = re.findall(r'<strong>[^<]+</strong>\s*·\s*([^<·]+?)(?=\s*·|\s*<|$)', prov_html)
    if strong_auction_matches:
        auction = strong_auction_matches[-1].strip()
    elif "·" in prov_html:
        # 回退：取最后一段
        parts = re.split(r'·', prov_html)
        if len(parts) > 1:
            auction = re.sub(r'<[^>]+>', '', parts[-2]).strip() if len(parts) >= 2 else ""

    # 清理卖家中的 <br>
    seller = re.sub(r'<br>.*$', '', seller, flags=re.DOTALL).strip()

    return seller, auction, lot, year


def extract_from_html(html, coin_id):
    """从 index.html 提取指定 ID 的卡片数据"""
    # 找到卡片注释
    pattern = re.compile(rf'<!-- {re.escape(coin_id)}([^>]*?) -->\s*(.*?)(?=(?:<!-- [A-Z]+-\d+)|</div><!-- end coin-grid -->)', re.DOTALL)
    m = pattern.search(html)
    if not m:
        return None

    block = m.group(0)

    def extract_text(selector_pattern):
        m = re.search(selector_pattern, block, re.DOTALL)
        if m:
            return re.sub(r'\s+', ' ', m.group(1)).strip()
        return ""

    # 提取各字段
    data = {
        "id": coin_id,
        "category": re.search(r'data-category="([^"]+)"', block).group(1) if re.search(r'data-category="([^"]+)"', block) else "",
        "grade": extract_text(r'<span class="coin-grade"[^>]*>([^<]+)</span>'),
        "title": extract_text(r'<h2 class="coin-title"[^>]*>([^<]+)</h2>'),
        "subtitle": extract_text(r'<p class="coin-subtitle"[^>]*>([^<]+)</p>'),
        "image": re.search(r'src="(images/[^"]+)"', block).group(1) if re.search(r'src="(images/[^"]+)"', block) else "",
        "era": "", "denomination": "", "material": "", "reference": "",
    }

    # 解析 detail grid（4 项：年代/面值/材质/参考）
    # 每个 detail-item 结构: <div class="detail-item"><div class="detail-label">X</div><div class="detail-value">Y</div></div>
    # 注意：detail-value 的闭合 </div> 与 item 的闭合 </div> 是相邻的，
    # 所以用 lookahead (?=</div>) 来消费它而不捕获，防止下一个 regex 用时找不到
    detail_items = re.findall(
        r'<div class="detail-item">'
        r'.*?<div class="detail-label"[^>]*>[^<]*</div>'
        r'.*?<div class="detail-value"[^>]*>([^<]*)</div>'
        r'(?=</div>)',
        block, re.DOTALL
    )
    labels = ["era", "denomination", "material", "reference"]
    for i, val in enumerate(detail_items[:4]):
        data[labels[i]] = re.sub(r'\s+', ' ', val).strip()

    # 描述
    desc_match = re.search(r'<div class="description">.*?<p[^>]*>(.*?)</p>', block, re.DOTALL)
    if desc_match:
        data["description"] = re.sub(r'\s+', ' ', desc_match.group(1)).strip()

    # obv/rev
    obv = re.search(r'<div class="obv"[^>]*>(.*?)</div>', block, re.DOTALL)
    if obv:
        data["obverse"] = re.sub(r'\s+', ' ', obv.group(1)).strip()
    rev = re.search(r'<div class="rev"[^>]*>(.*?)</div>', block, re.DOTALL)
    if rev:
        data["reverse"] = re.sub(r'\s+', ' ', rev.group(1)).strip()

    # tags
    tags = re.findall(r'<span class="tag"[^>]*>([^<]+)</span>', block)
    data["tags"] = tags

    # price
    price = re.search(r'<div class="price-value"[^>]*>([^<]+)</div>', block)
    if price:
        price_str = price.group(1).strip()
        code, amount, cny = parse_price(price_str)
        data["price"] = price_str
        data["price_currency"] = code or ""
        data["price_cny"] = cny if cny is not None else ""

    # provenance
    prov = re.search(r'<div class="prov-compact"[^>]*>(.*?)</div>', block, re.DOTALL)
    if prov:
        prov_str = prov.group(1)
        seller, auction, lot, year = parse_provenance(prov_str)
        data["seller"] = seller
        data["auction"] = auction
        data["lot"] = lot
        data["year"] = year
        data["provenance_raw"] = re.sub(r'\s+', ' ', prov_str).strip()

    return data


def save_coin(coin_id, confirmed=False, git_commit=None):
    """保存单枚钱币到 JSON + Markdown + Excel"""
    ensure_archive_dir()

    if not os.path.exists(INDEX_PATH):
        print(f"❌ {INDEX_PATH} 不存在")
        return False

    with open(INDEX_PATH, "r", encoding="utf-8-sig", newline="\n") as f:
        html = f.read()

    data = extract_from_html(html, coin_id)
    if not data:
        print(f"❌ 在 index.html 中未找到 {coin_id}")
        return False

    # 添加时间戳
    now = datetime.now().isoformat(timespec="seconds")
    if confirmed:
        data["confirmed_at"] = now
    else:
        data["created_at"] = now
    if git_commit:
        data["git_commit"] = git_commit

    # 图片信息
    if data["image"]:
        img_path = os.path.join(BASE_DIR, data["image"])
        if os.path.exists(img_path):
            data["image_size"] = os.path.getsize(img_path)

    # 1. 写 JSON
    json_path = os.path.join(ARCHIVE_DIR, f"{coin_id}.json")
    with open(json_path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {json_path}")

    # 2. 写 Markdown
    md_path = os.path.join(ARCHIVE_DIR, f"{coin_id}.md")
    md = build_markdown(data)
    with open(md_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(md)
    print(f"  ✓ {md_path}")

    # 3. 更新 Excel
    update_excel(data)
    print(f"  ✓ {EXCEL_PATH}")

    return True


def build_markdown(d):
    """生成单枚钱币的 Markdown 描述"""
    tags_str = " ".join([f"`{t}`" for t in d.get("tags", [])])
    lines = [
        f"# {d['id']} — {d.get('title', '')}",
        "",
        f"**{d.get('subtitle', '')}**",
        "",
        f"![{d.get('title', '')}]({d.get('image', '')})",
        "",
        "## 基本信息",
        "",
        f"| 字段 | 内容 |",
        f"|------|------|",
        f"| 年代 | {d.get('era', '')} |",
        f"| 面值 | {d.get('denomination', '')} |",
        f"| 材质 | {d.get('material', '')} |",
        f"| 品相 | {d.get('grade', '')} |",
        f"| 参考 | {d.get('reference', '')} |",
        f"| 分类 | {d.get('category', '')} |",
        "",
        "## 历史背景",
        "",
        d.get('description', ''),
        "",
        "## 正面",
        "",
        d.get('obverse', ''),
        "",
        "## 背面",
        "",
        d.get('reverse', ''),
        "",
        "## 收藏信息",
        "",
        f"| 字段 | 内容 |",
        f"|------|------|",
        f"| 价格 | {d.get('price', '')} |",
        f"| 折合人民币 | ¥{d.get('price_cny', '')} |" if d.get('price_cny') else "| 价格折算 | — |",
        f"| 卖家 | {d.get('seller', '')} |",
        f"| 拍卖 | {d.get('auction', '')} |",
        f"| Lot | {d.get('lot', '')} |",
        f"| 年份 | {d.get('year', '')} |",
        "",
        "## 标签",
        "",
        tags_str,
        "",
        "## 元数据",
        "",
        f"- 创建时间: {d.get('created_at', '')}",
        f"- 确认时间: {d.get('confirmed_at', '')}",
        f"- Git commit: `{d.get('git_commit', '') or '未记录'}`",
        f"- 图片: {d.get('image', '')} ({d.get('image_size', 0)} bytes)",
        "",
    ]
    return "\n".join(lines)


def update_excel(data):
    """追加或更新 Excel 行（按 ID 唯一）"""
    from openpyxl import Workbook, load_workbook

    headers = [
        "ID", "国家/朝代", "标题", "副标题", "年代", "面值", "材质",
        "品相", "参考", "分类", "正面", "背面", "标签",
        "价格（原币种）", "币种", "折合人民币", "卖家", "拍卖", "Lot", "年份",
        "图片路径", "图片大小(B)", "历史背景",
        "创建时间", "确认时间", "Git commit",
    ]

    # 读取现有 Excel
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        # 找到已存在 ID 的行
        existing_row = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == data["id"]:
                existing_row = row_idx
                break
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "钱币目录"
        ws.append(headers)
        existing_row = None

    # 行数据
    tags_str = ", ".join(data.get("tags", []))
    row = [
        data.get("id", ""),
        data.get("subtitle", "").split("·")[0].strip() if data.get("subtitle") else "",
        data.get("title", ""),
        data.get("subtitle", ""),
        data.get("era", ""),
        data.get("denomination", ""),
        data.get("material", ""),
        data.get("grade", ""),
        data.get("reference", ""),
        data.get("category", ""),
        data.get("obverse", ""),
        data.get("reverse", ""),
        tags_str,
        data.get("price", ""),
        data.get("price_currency", ""),
        data.get("price_cny", ""),
        data.get("seller", ""),
        data.get("auction", ""),
        data.get("lot", ""),
        data.get("year", ""),
        data.get("image", ""),
        data.get("image_size", ""),
        data.get("description", ""),
        data.get("created_at", ""),
        data.get("confirmed_at", ""),
        data.get("git_commit", ""),
    ]

    if existing_row:
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=existing_row, column=col_idx, value=val)
    else:
        ws.append(row)

    # 美化：设置列宽
    col_widths = {
        1: 10, 2: 18, 3: 30, 4: 28, 5: 15, 6: 18, 7: 22, 8: 14, 9: 22,
        10: 10, 11: 30, 12: 30, 13: 30, 14: 12, 15: 8, 16: 12, 17: 18,
        18: 18, 19: 8, 20: 8, 21: 20, 22: 10, 23: 60, 24: 20, 25: 20, 26: 12,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{col}"].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)


def list_archived():
    """列出已存档的所有钱币"""
    if not os.path.exists(ARCHIVE_DIR):
        print("存档目录不存在")
        return
    files = sorted([f for f in os.listdir(ARCHIVE_DIR) if f.endswith(".json")])
    print(f"已存档 {len(files)} 枚钱币：")
    for f in files:
        coin_id = f.replace(".json", "")
        with open(os.path.join(ARCHIVE_DIR, f), "r", encoding="utf-8") as fp:
            d = json.load(fp)
        print(f"  - {coin_id}: {d.get('title', '')} ({d.get('era', '')})")


def rebuild_from_html():
    """从 index.html 重建所有存档（一次性操作）"""
    if not os.path.exists(INDEX_PATH):
        print("❌ index.html 不存在")
        return
    with open(INDEX_PATH, "r", encoding="utf-8-sig", newline="\n") as f:
        html = f.read()
    # 找所有卡片注释
    ids = re.findall(r'<!-- ([A-Z]+-\d+)\b', html)
    # 特殊处理 ROM-004 批量组
    if "ROM-004 批量组" in html:
        ids.append("ROM-004")
    # 去重保序
    seen = set()
    unique_ids = []
    for i in ids:
        if i not in seen:
            seen.add(i)
            unique_ids.append(i)
    print(f"找到 {len(unique_ids)} 张卡片，正在存档...")
    for cid in unique_ids:
        if save_coin(cid):
            print(f"  ✓ {cid}")
        else:
            print(f"  ✗ {cid} 跳过")
    print(f"\n完成。存档位置: {ARCHIVE_DIR}")


def get_git_commit():
    """获取当前 HEAD commit SHA（前 7 位）"""
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=BASE_DIR,
            capture_output=True, text=True, timeout=5
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass
    return None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法：")
        print("  python save_coin.py {ID}         # 保存单枚（标记创建）")
        print("  python save_coin.py {ID} --confirm  # 保存单枚（标记确认）")
        print("  python save_coin.py --list       # 列出所有已存档")
        print("  python save_coin.py --rebuild    # 从 index.html 重建所有存档")
        sys.exit(0)

    if sys.argv[1] == "--list":
        list_archived()
    elif sys.argv[1] == "--rebuild":
        rebuild_from_html()
    else:
        coin_id = sys.argv[1]
        confirm = "--confirm" in sys.argv
        git_sha = get_git_commit()
        save_coin(coin_id, confirmed=confirm, git_commit=git_sha)
